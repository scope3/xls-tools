import openpyxl
from .xlrd_like import XlrdSheetLike, XlrdCellLike, XlrdWorkbookLike


class OpenpyxlSheetLike(XlrdSheetLike):

    @property
    def xlsx(self):
        """
        Allow native access to openpyxl object
        :return:
        """
        return self._xlsx

    def __init__(self, xlsx_sheet):
        """
        We have to access the openpyxl sheet's internal dict of _cells in order to filter out Nones when computing
        the spreadsheet's range
        :param xlsx_sheet:
        """
        self._xlsx = xlsx_sheet
        max_row = max_col = 0
        for coord, cell in xlsx_sheet._cells.items():  # we have to do this because openpyxl is.. designed for purposes different from mine
            if cell.data_type == 'n':
                continue
            max_row = max([max_row, coord[0]])
            max_col = max([max_col, coord[1]])
        self._nrows = max_row
        self._ncols = max_col

    @property
    def name(self):
        return self._xlsx.title

    @property
    def ncols(self):
        return self._ncols

    @property
    def nrows(self):
        return self._nrows

    def row(self, row):
        """
        zero-indexed!
        :param row:
        :return:
        """
        row += 1
        if row > self._nrows:
            raise IndexError

        rows = list(self._xlsx.iter_rows(min_row=row, max_row=row, max_col=self._ncols))  # 2nd order list)
        return [XlrdCellLike(k.value) for k in rows[0]]

    def get_rows(self):
        rows = list(self._xlsx.iter_rows(min_row=1, max_row=self._nrows, max_col=self._ncols))
        for row in rows:
            yield [XlrdCellLike(k.value) for k in row]

    def col(self, col):
        """
        Zero-indexed!
        AND, per xlrd, negative index is counting from ncols
        :param col:
        :return:
        """
        if col < 0:
            col = self.ncols + col + 1
        else:
            col += 1
        if col > self._ncols:
            raise IndexError

        cols = list(self._xlsx.iter_cols(min_col=col, max_col=col, max_row=self._nrows))
        return [XlrdCellLike(k.value) for k in cols[0]]

    def cell(self, row, col):
        row += 1
        col += 1
        if row > self._nrows or col > self._ncols:
            raise IndexError

        cell = self._xlsx.cell(row, col)
        return XlrdCellLike(cell.value)


class OpenpyXlrdWorkbook(XlrdWorkbookLike):

    @classmethod
    def from_file(cls, file, **kwargs):
        return cls(openpyxl.load_workbook(file, **kwargs))

    @property
    def book(self):
        """
        Allow native access to book
        :return:
        """
        return self._book

    def __init__(self, xl_book):
        """
        :param xl_book: an initialized Openpyxl Workbook
        """
        self._book = xl_book
        self._names = {k: i for i, k in enumerate(self._book.sheetnames)}
        self._sheets = [OpenpyxlSheetLike(self._book[k]) for k in self._book.sheetnames]

    def sheet_names(self):
        return self._book.sheetnames

    def sheet_by_name(self, name):
        return self._sheets[self._names[name]]

    def sheet_by_index(self, index):
        return self._sheets[index]

    def sheets(self):
        return self._sheets

    @property
    def filename(self):
        return self._book.file
